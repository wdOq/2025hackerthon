import csv
import hashlib
import io
import os
import zipfile
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import requests
from bs4 import BeautifulSoup

# Add parent directory to path to import utils
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.error_handler import create_error_result


def _norm_cas(s: str) -> str:
    if not s:
        return ""
    s = s.strip()
    # Uppercase and remove spaces
    s = s.replace(" ", "").upper()
    # Some files may have CAS in the form XXXXXXX without dashes; try to insert
    digits = [c for c in s if c.isdigit()]
    if len(digits) >= 5 and "-" not in s:
        # Last digit is check, preceding two are middle, rest is first
        chk = digits[-1]
        mid = "".join(digits[-3:-1])
        first = "".join(digits[:-3])
        if first and mid:
            return f"{int(first)}-{mid}-{chk}"
    return s


class USTscaInventory:


    def __init__(self, name: str, jurisdiction: str, slug: str,
                 inventory_csv: str, queries_csv: Optional[str] = None,
                 dump_all: bool = False, fields: Optional[List[str]] = None):
        self.name = name
        self.jurisdiction = jurisdiction
        self.slug = slug
        self.inventory_csv = inventory_csv
        self.queries_csv = queries_csv
        self.dump_all = dump_all
        self.fields = fields
        # For compatibility with regwatch (use real URL if provided)
        if str(inventory_csv).lower().startswith(("http://", "https://")):
            self.url = str(inventory_csv)
        else:
            self.url = f"file://{inventory_csv}"

    def _ensure_local_file(self, src: str, target_dir: Path, default_name: str) -> Path:
        """Ensure source exists locally; download if URL.
        Supports: direct CSV/XLSX/ZIP, or an EPA landing page that links to a ZIP.
        """
        target_dir.mkdir(parents=True, exist_ok=True)
        # URL source
        if src.lower().startswith(("http://", "https://")):
            # If this looks like a landing page (not a data file), discover the ZIP link
            if not any(src.lower().endswith(ext) for ext in ('.csv', '.xlsx', '.zip')):
                try:
                    resp = requests.get(src, timeout=20)
                    resp.raise_for_status()
                    soup = BeautifulSoup(resp.text, 'html.parser')
                    anchors = soup.find_all('a', href=True)
                    # prefer CSV Inventory zip
                    candidates: List[str] = []
                    for a in anchors:
                        href = a['href']
                        text = (a.get_text(" ", strip=True) or "").lower()
                        if not href.lower().endswith('.zip'):
                            continue
                        score = 0
                        if 'csv' in text:
                            score += 3
                        if 'inventory' in text or 'tsca' in text:
                            score += 2
                        if 'csv' in href.lower():
                            score += 1
                        # encode score in front for simple sorting
                        candidates.append(f"{score:02d}|{href}")
                    if candidates:
                        candidates.sort(reverse=True)
                        best = candidates[0].split('|',1)[1]
                        from urllib.parse import urljoin
                        src = urljoin(resp.url, best)
                except Exception:
                    pass

            # Guess filename from src
            fname = os.path.basename(src) or default_name
            dl_path = target_dir / fname
            # Download if not cached
            if not dl_path.exists() or dl_path.stat().st_size == 0:
                resp = requests.get(src, timeout=20)
                resp.raise_for_status()
                dl_path.write_bytes(resp.content)
            # If ZIP: choose TSCAINV* CSV else XLSX
            if dl_path.suffix.lower() == ".zip":
                with zipfile.ZipFile(io.BytesIO(dl_path.read_bytes())) as zf:
                    names = zf.namelist()
                    def pick(members: List[str], exts: Tuple[str, ...]) -> Optional[str]:
                        cands = [m for m in members if m.lower().endswith(exts)]
                        inv = [m for m in cands if 'tscainv' in m.lower() or 'inventory' in m.lower()]
                        return inv[0] if inv else (cands[0] if cands else None)
                    target = pick(names, ('.csv',)) or pick(names, ('.xlsx',))
                    if not target:
                        raise RuntimeError("ZIP missing TSCAINV CSV/XLSX")
                    out_path = target_dir / Path(target).name
                    with zf.open(target) as srcf, open(out_path, 'wb') as outf:
                        outf.write(srcf.read())
                    return out_path
            return dl_path
        # Local path
        p = Path(src)
        if not p.exists():
            raise FileNotFoundError(f"File not found: {src}")
        return p

    def _load_inventory(self) -> Tuple[Dict[str, Dict[str, str]], List[str], List[Dict[str, str]], str]:
        # Ensure local data file (download if URL or ZIP). May return CSV or XLSX.
        inv_path = self._ensure_local_file(self.inventory_csv, Path('data'), 'tsca_inventory.csv')
        suffix = inv_path.suffix.lower()
        index: Dict[str, Dict[str, str]] = {}
        headers: List[str] = []
        rows_list: List[Dict[str, str]] = []
        cas_col_detected: str = ''
        if suffix == '.csv':
            with inv_path.open('r', encoding='utf-8', errors='ignore', newline='') as f:
                reader = csv.DictReader(f)
                headers = [h.strip() for h in (reader.fieldnames or [])]
                # Flexible CAS header detection
                import re as _re
                def _is_cas_header(h: str) -> bool:
                    hn = _re.sub(r"[^a-z0-9]", "", h.lower())
                    return ("cas" in hn) and ("number" in hn or "rn" in hn or "registry" in hn or "regno" in hn or "reg" in hn)
                cas_candidates = [h for h in headers if _is_cas_header(h)]
                if not cas_candidates:
                    raise RuntimeError(f"Inventory CSV must contain a CAS column (e.g., CASRN or CAS Number); headers={headers}")
                cas_col = cas_candidates[0]
                cas_col_detected = cas_col
                for row in reader:
                    cas_raw = row.get(cas_col, '')
                    cas = _norm_cas(cas_raw)
                    if not cas:
                        continue
                    rows_list.append(row)
                    index[cas] = row
            return index, headers, rows_list, cas_col_detected
        elif suffix == '.xlsx':
            # Parse XLSX via openpyxl
            try:
                from openpyxl import load_workbook  # type: ignore
            except Exception as e:
                raise RuntimeError("Inventory is XLSX; please install openpyxl to parse it") from e
            wb = load_workbook(filename=str(inv_path), read_only=True, data_only=True)
            # Pick the first worksheet that has a CAS column in the first 20 rows
            ws = None
            cas_idx = None
            for sheet in wb.worksheets:
                rows_iter = sheet.iter_rows(values_only=True)
                probe_rows = []
                for _ in range(20):
                    try:
                        probe_rows.append(next(rows_iter))
                    except StopIteration:
                        break
                # Try to find header row
                header_row = None
                for r in probe_rows:
                    if not r:
                        continue
                    cells = [str(c).strip() if c is not None else '' for c in r]
                    lows = [c.lower() for c in cells]
                    for i, h in enumerate(lows):
                        hn = _re.sub(r"[^a-z0-9]", "", h)
                        if ("cas" in hn) and ("number" in hn or "rn" in hn or "registry" in hn or "regno" in hn or "reg" in hn):
                            header_row = cells
                            cas_idx = i
                            break
                    if header_row is not None:
                        break
                if header_row is not None:
                    ws = sheet
                    headers = header_row
                    break
            if ws is None or cas_idx is None:
                raise RuntimeError("Inventory XLSX must contain a CAS column (e.g., CASRN or CAS Number)")

            # Build header index map (lowercased)
            header_index = {i: (h.lower()) for i, h in enumerate(headers)}
            cas_col_detected = headers[cas_idx]
            # Iterate remaining rows starting after the detected header
            start_collecting = False
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() if c is not None else '' for c in row]
                if not start_collecting:
                    # Wait until we hit the header row again (exact match) to start
                    if [c.lower() for c in cells] == [h.lower() for h in headers]:
                        start_collecting = True
                    continue
                if not cells:
                    continue
                cas_raw = cells[cas_idx] if cas_idx < len(cells) else ''
                cas = _norm_cas(cas_raw)
                if not cas:
                    continue
                record = {}
                for i, h in enumerate(headers):
                    record[h] = '' if i >= len(cells) or cells[i] is None else str(cells[i])
                rows_list.append(record)
                index[cas] = record
            if cas_idx is None:
                raise RuntimeError("CAS column not detected after scanning header rows in XLSX")
            return index, headers, rows_list, cas_col_detected
        else:
            raise RuntimeError(f"Unsupported inventory format: {inv_path}")

    def _load_queries(self) -> List[Dict[str, str]]:
        # Allow queries_csv to be optional. If not provided or not found, return empty list.
        if not self.queries_csv:
            return []
        try:
            q_path = self._ensure_local_file(self.queries_csv, Path('inputs'), 'chemicals.csv')
        except FileNotFoundError:
            return []
        with q_path.open('r', encoding='utf-8', errors='ignore', newline='') as f:
            reader = csv.DictReader(f)
            items = []
            for row in reader:
                cas = _norm_cas(row.get('CAS', '') or row.get('cas', '') or row.get('Cas', ''))
                name = row.get('Name') or row.get('name') or row.get('ChemicalName') or ''
                if cas or name:
                    items.append({'CAS': cas, 'Name': name})
        return items

    def fetch(self):
        try:
            # Ensure local data file and parse
            inventory, headers, all_rows, cas_col = self._load_inventory()
            queries = self._load_queries()
        except Exception as e:
            return create_error_result(
                self.name,
                f"無法載入 inventory 或 queries 檔案：{e}",
                category="chemical_inventory"
            )

        # If no queries provided, just emit a brief catalog of N records (no per-section lines)
        per_records: List[Dict[str, str]] = []
        if self.dump_all:
            name_candidates = ['ChemName;DEF', 'ChemName', 'GenericName', 'CA Index Name', 'Chemical Name']
            fields = self.fields or ['CASRN', 'casregno', 'ChemName;DEF', 'ChemName', 'GenericName', 'UVCB', 'FLAG', 'ACTIVITY', 'UID', 'EXP']
            for row in all_rows:
                # Build a record that uses original CSV/XLSX column names as keys
                rec = {}
                # Always include a normalized CAS under CASRN if available column exists; else under casregno
                if cas_col and cas_col in row:
                    rec[cas_col] = _norm_cas(row.get(cas_col, ''))
                # Add selected fields if present
                for k in fields:
                    if k in row and row.get(k):
                        rec[k] = row.get(k)
                # Fallback name heading for convenience (non-critical)
                name = next((row.get(k) for k in name_candidates if k in row and row.get(k)), '')
                # Keep a combined text for quick grep
                text = "; ".join([f"{k}={v}" for k, v in rec.items()])
                rec.setdefault('text', text)
                rec.setdefault('length', len(text))
                per_records.append(rec)
        elif queries:
            for q in queries:
                cas = q.get('CAS', '')
                name = q.get('Name', '')
                match = inventory.get(cas)
                listed = match is not None
                if listed:
                    details = {k: v for k, v in (match or {}).items() if v and k}
                    keys_prefer = ['CASRN', 'CAS Number', 'casregno', 'CA Index Name', 'Chemical Name', 'ChemName', 'Active/Inactive', 'Status']
                    ordered = []
                    for k in keys_prefer:
                        if k in details:
                            ordered.append(f"{k}={details[k]}")
                    text = f"listed=true; " + ", ".join(ordered) if ordered else "listed=true"
                else:
                    text = "listed=false"
                per_records.append({
                    'part': '',
                    'section_citation': cas,
                    'section_heading': name or cas,
                    'text': text,
                    'length': len(text),
                })

        combined_text = "\n".join([r.get('text','') for r in per_records]) if per_records else ""
        sha256 = hashlib.sha256((combined_text or 'empty').encode('utf-8')).hexdigest()

        display_title = "TSCA Inventory (CAS-based)" if cas_col else "TSCA PMN/ACC New Chemicals"
        return {
            'title': display_title,
            'version_date': None,
            'regulation_number': 'TSCA Inventory',
            'sources': [{'inventory_source': self.url, 'records_indexed': len(inventory)}],
            'structured_sections': None,
            'full_content': combined_text or None,
            'content_length': len(combined_text),
            'excerpt': (combined_text or '')[:1000],
            'sha256': sha256,
            'notes': f"TSCA Inventory ready; queries={len(queries)}; indexed={len(inventory)}; dump_all={self.dump_all}",
            'per_section_records': per_records,
        }
